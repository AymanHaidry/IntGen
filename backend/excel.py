"""Native, editable Excel chart export for IntGen using openpyxl."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import (
    AreaChart,
    BarChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
    Series,
)
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Font, PatternFill, Alignment

DEFAULT_PALETTE = [
    "CC785C", "6A8EAE", "788C5D", "D4A27F",
    "BF4D43", "5C6B73", "A87C5F", "8DA9C4",
]

HEADER_FILL = PatternFill("solid", fgColor="2B2A27")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Georgia")
TITLE_FONT = Font(size=14, bold=True, name="Georgia", color="2B2A27")


def _hex(color: str | None, fallback: str) -> str:
    if not color:
        return fallback
    return color.lstrip("#").upper()


def _solid_props(hex_color: str) -> GraphicalProperties:
    props = GraphicalProperties()
    props.solidFill = hex_color
    return props


def build_xlsx(cfg: dict[str, Any]) -> bytes:
    chart_type = cfg.get("chart_type", "column")
    title = cfg.get("title") or "IntGen Chart"
    categories = [str(c) for c in cfg.get("categories", [])]
    series = cfg.get("series", []) or []
    bar_colors = cfg.get("bar_colors")

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # ---- Write the data table ----
    # Row 1: title. Row 3: header. Row 4+: data.
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT

    header_row = 3
    ws.cell(row=header_row, column=1, value="Category")
    for j, s in enumerate(series):
        c = ws.cell(row=header_row, column=2 + j, value=s.get("name") or f"Series {j + 1}")
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=header_row, column=1).fill = HEADER_FILL
    ws.cell(row=header_row, column=1).font = HEADER_FONT

    for i, cat in enumerate(categories):
        ws.cell(row=header_row + 1 + i, column=1, value=cat)
        for j, s in enumerate(series):
            vals = s.get("values", [])
            v = vals[i] if i < len(vals) else None
            ws.cell(row=header_row + 1 + i, column=2 + j, value=v)

    ws.column_dimensions["A"].width = 20
    for j in range(len(series)):
        ws.column_dimensions[chr(66 + j)].width = 16

    n_rows = len(categories)
    last_data_row = header_row + n_rows

    # ---- Build the native chart ----
    chart = _make_chart(chart_type)
    chart.title = title
    chart.height = 10
    chart.width = 18
    chart.style = 2

    if isinstance(chart, ScatterChart):
        _populate_scatter(chart, ws, series, header_row, last_data_row)
    else:
        data = Reference(
            ws,
            min_col=2,
            max_col=1 + len(series),
            min_row=header_row,
            max_row=last_data_row,
        )
        cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_data_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Per-point coloring for single-series bar / pie / column charts.
        single = len(series) == 1
        if single and isinstance(chart, (BarChart, PieChart)):
            plot_series = chart.series[0]
            for i in range(n_rows):
                override = bar_colors[i] if bar_colors and i < len(bar_colors) else None
                color = _hex(override, DEFAULT_PALETTE[i % len(DEFAULT_PALETTE)])
                dp = DataPoint(idx=i)
                dp.graphicalProperties = _solid_props(color)
                plot_series.data_points.append(dp)
        else:
            # Color whole series from the palette / explicit series color.
            for idx, s in enumerate(chart.series):
                override = series[idx].get("color") if idx < len(series) else None
                color = _hex(override, DEFAULT_PALETTE[idx % len(DEFAULT_PALETTE)])
                s.graphicalProperties = _solid_props(color)
                if isinstance(chart, LineChart):
                    s.graphicalProperties.line.solidFill = color
                    s.graphicalProperties.line.width = 28000
                    s.marker = Marker(symbol="circle", size=6)

    ws.add_chart(chart, f"{chr(66 + len(series) + 1)}3")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _make_chart(chart_type: str):
    if chart_type in ("column", "bar", "stacked_bar"):
        chart = BarChart()
        if chart_type == "bar":
            chart.type = "bar"  # horizontal
        else:
            chart.type = "col"
        if chart_type == "stacked_bar":
            chart.grouping = "stacked"
            chart.overlap = 100
        return chart
    if chart_type == "line":
        return LineChart()
    if chart_type == "area":
        return AreaChart()
    if chart_type in ("pie", "donut"):
        return PieChart()
    if chart_type == "scatter":
        return ScatterChart()
    # default
    return BarChart()


def _populate_scatter(chart, ws, series, header_row, last_data_row):
    x_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_data_row)
    for j, s in enumerate(series):
        y_ref = Reference(
            ws, min_col=2 + j, min_row=header_row, max_row=last_data_row
        )
        ser = Series(y_ref, x_ref, title_from_data=True)
        ser.marker = Marker(symbol="circle", size=7)
        ser.graphicalProperties.line.noFill = True
        chart.series.append(ser)
    chart.x_axis.title = "X"
    chart.y_axis.title = "Y"
